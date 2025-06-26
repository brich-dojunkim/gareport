import json
import pandas as pd
from datetime import datetime
from google.analytics.data_v1beta import BetaAnalyticsDataClient
from google.analytics.data_v1beta.types import (
    DateRange, Dimension, Metric, RunReportRequest, OrderBy, FilterExpression, Filter
)
from google.oauth2 import service_account
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class UserSignupPathAnalyzer:
    def __init__(self, credentials_path, property_id):
        """실제 유저 기준 가입 경로 분석기"""
        self.property_id = property_id
        
        # GA4 클라이언트 초기화
        credentials = service_account.Credentials.from_service_account_file(
            credentials_path,
            scopes=['https://www.googleapis.com/auth/analytics.readonly']
        )
        self.client = BetaAnalyticsDataClient(credentials=credentials)
    
    def analyze_signup_sources(self, days=30):
        """실제 가입자의 소스별 분석"""
        print("🔍 실제 가입자의 트래픽 소스 분석 중...")
        
        # 1. 가입 완료 유저의 세션 소스 분석
        signup_sources = self._get_signup_user_sources(days)
        
        # 2. 전체 방문자의 소스 분석 (신규+재방문 모두 포함)
        all_visitor_sources = self._get_all_visitor_sources(days)
        
        # 3. 가입 페이지 방문자의 소스 분석 (제거 - 의미없는 지표)
        # signup_page_sources = self._get_signup_page_visitors_only(days)
        
        # 4. 트래픽 품질 지표 수집 (제거 - 의미없는 지표)
        # traffic_quality = self._get_traffic_quality_metrics(days)
        
        # 4. 종합 분석 (간소화된 버전)
        analysis_result = self._analyze_conversion_paths(
            signup_sources, all_visitor_sources
        )
        
        # 5. 엑셀 리포트 생성
        excel_filename = self._create_excel_report(analysis_result)
        
        return analysis_result
    
    def _get_signup_user_sources(self, days):
        """가입 완료한 실제 유저들의 세션 소스 분석"""
        print("  📊 가입 완료 유저의 트래픽 소스 수집...")
        
        request = RunReportRequest(
            property=f"properties/{self.property_id}",
            dimensions=[
                Dimension(name="sessionSource"),
                Dimension(name="sessionMedium"),
            ],
            metrics=[
                Metric(name="activeUsers"),  # 실제 고유 사용자 수
                Metric(name="eventCount"),
                Metric(name="sessions"),
            ],
            date_ranges=[DateRange(start_date=f"{days}daysAgo", end_date="today")],
            dimension_filter=FilterExpression(
                filter=Filter(
                    field_name="eventName",
                    string_filter=Filter.StringFilter(
                        match_type=Filter.StringFilter.MatchType.EXACT,
                        value="회원가입2"
                    )
                )
            ),
            order_bys=[OrderBy(metric=OrderBy.MetricOrderBy(metric_name="activeUsers"), desc=True)],
            limit=100
        )
        
        response = self.client.run_report(request=request)
        return self._response_to_dataframe(response)
    
    def _get_all_visitor_sources(self, days):
        """전체 방문자들의 세션 소스 분석 (first_visit 대신 전체 activeUsers 사용)"""
        print("  🌐 전체 방문자의 트래픽 소스 수집...")
        
        request = RunReportRequest(
            property=f"properties/{self.property_id}",
            dimensions=[
                Dimension(name="sessionSource"),
                Dimension(name="sessionMedium"),
            ],
            metrics=[
                Metric(name="activeUsers"),  # 전체 고유 사용자
                Metric(name="newUsers"),     # 신규 사용자
                Metric(name="sessions"),
            ],
            date_ranges=[DateRange(start_date=f"{days}daysAgo", end_date="today")],
            order_bys=[OrderBy(metric=OrderBy.MetricOrderBy(metric_name="activeUsers"), desc=True)],
            limit=100
        )
        
        response = self.client.run_report(request=request)
        return self._response_to_dataframe(response)
    
    def _get_signup_page_visitors_only(self, days):
        """가입 페이지 방문자만 수집 (가입 이벤트 필터 없이)"""
        print("  📝 가입 페이지 방문자의 트래픽 소스 수집...")
        
        request = RunReportRequest(
            property=f"properties/{self.property_id}",
            dimensions=[
                Dimension(name="sessionSource"),
                Dimension(name="sessionMedium"),
            ],
            metrics=[
                Metric(name="activeUsers"),
                Metric(name="sessions"),
                Metric(name="screenPageViews"),
            ],
            date_ranges=[DateRange(start_date=f"{days}daysAgo", end_date="today")],
            dimension_filter=FilterExpression(
                filter=Filter(
                    field_name="pagePath",
                    string_filter=Filter.StringFilter(
                        match_type=Filter.StringFilter.MatchType.CONTAINS,
                        value="provider-join"
                    )
                )
            ),
            order_bys=[OrderBy(metric=OrderBy.MetricOrderBy(metric_name="activeUsers"), desc=True)],
            limit=100
        )
        
        response = self.client.run_report(request=request)
        return self._response_to_dataframe(response)
    
    def _get_traffic_quality_metrics(self, days):
        """트래픽 품질 지표 수집 (참여율, 이탈률, 세션시간 등)"""
        print("  📊 트래픽 품질 지표 수집...")
        
        request = RunReportRequest(
            property=f"properties/{self.property_id}",
            dimensions=[
                Dimension(name="sessionSource"),
                Dimension(name="sessionMedium"),
            ],
            metrics=[
                Metric(name="sessions"),
                Metric(name="newUsers"),
                Metric(name="engagementRate"),
                Metric(name="bounceRate"),
                Metric(name="averageSessionDuration"),
                Metric(name="screenPageViewsPerSession"),
            ],
            date_ranges=[DateRange(start_date=f"{days}daysAgo", end_date="today")],
            order_bys=[OrderBy(metric=OrderBy.MetricOrderBy(metric_name="newUsers"), desc=True)],
            limit=100
        )
        
        response = self.client.run_report(request=request)
        return self._response_to_dataframe(response)
    
    def _get_signup_user_journey(self, days):
        """가입자들의 사용자 여정 분석"""
        print("  🛤️ 가입자들의 사용자 여정 분석...")
        
        request = RunReportRequest(
            property=f"properties/{self.property_id}",
            dimensions=[
                Dimension(name="sessionSource"),
                Dimension(name="sessionMedium"),
                Dimension(name="landingPage"),
                Dimension(name="pagePath"),
            ],
            metrics=[
                Metric(name="activeUsers"),
                Metric(name="sessions"),
            ],
            date_ranges=[DateRange(start_date=f"{days}daysAgo", end_date="today")],
            dimension_filter=FilterExpression(
                filter=Filter(
                    field_name="eventName",
                    string_filter=Filter.StringFilter(
                        match_type=Filter.StringFilter.MatchType.EXACT,
                        value="회원가입2"
                    )
                )
            ),
            order_bys=[OrderBy(metric=OrderBy.MetricOrderBy(metric_name="activeUsers"), desc=True)],
            limit=200
        )
        
        response = self.client.run_report(request=request)
        return self._response_to_dataframe(response)
    
    def _response_to_dataframe(self, response):
        """GA4 응답을 DataFrame으로 변환"""
        if not response.rows:
            return pd.DataFrame()
        
        # 헤더 정보 추출
        dimension_headers = [header.name for header in response.dimension_headers]
        metric_headers = [header.name for header in response.metric_headers]
        
        # 데이터 변환
        data = []
        for row in response.rows:
            row_data = {}
            
            # 차원 데이터
            for i, dimension_value in enumerate(row.dimension_values):
                row_data[dimension_headers[i]] = dimension_value.value
            
            # 측정항목 데이터 (숫자 변환)
            for i, metric_value in enumerate(row.metric_values):
                try:
                    row_data[metric_headers[i]] = float(metric_value.value)
                except ValueError:
                    row_data[metric_headers[i]] = metric_value.value
            
            data.append(row_data)
        
        return pd.DataFrame(data)
    
    def _analyze_conversion_paths(self, signup_sources, all_visitor_sources):
        """전환 경로 종합 분석 (전체 방문자 기준)"""
        print("  🎯 핵심 전환 지표 분석...")
        
        analysis = {
            "core_funnel_analysis": []
        }
        
        # 소스별 핵심 퍼널 분석
        if not signup_sources.empty and not all_visitor_sources.empty:
            
            # 각 DataFrame에 source_medium 컬럼 추가
            signup_sources['source_medium'] = signup_sources['sessionSource'] + ' / ' + signup_sources['sessionMedium']
            all_visitor_sources['source_medium'] = all_visitor_sources['sessionSource'] + ' / ' + all_visitor_sources['sessionMedium']
            
            # 각 소스별로 데이터 집계
            signup_summary = signup_sources.groupby('source_medium').agg({
                'activeUsers': 'sum',
                'eventCount': 'sum',
                'sessions': 'sum'
            }).reset_index()
            
            all_visitor_summary = all_visitor_sources.groupby('source_medium').agg({
                'activeUsers': 'sum',  # 전체 방문자
                'newUsers': 'sum',     # 신규 방문자
                'sessions': 'sum'
            }).reset_index()
            
            # 전체 가입자 수 계산 (비율 계산용)
            total_signups = signup_summary['activeUsers'].sum()
            
            # 핵심 퍼널 데이터 생성
            core_data = []
            
            # 가입자가 있는 소스들을 기준으로 분석
            for _, signup_row in signup_summary.iterrows():
                source = signup_row['source_medium']
                signup_users = int(signup_row['activeUsers'])
                
                # 전체 방문자 데이터 찾기
                all_visitor_row = all_visitor_summary[all_visitor_summary['source_medium'] == source]
                all_visitors = int(all_visitor_row['activeUsers'].iloc[0]) if not all_visitor_row.empty else 0
                
                # 핵심 지표 계산
                conversion_rate = round((signup_users / all_visitors) * 100, 1) if all_visitors > 0 else 0
                contribution = round((signup_users / total_signups) * 100, 1) if total_signups > 0 else 0
                
                core_data.append({
                    'source_medium': source,
                    'all_visitors': all_visitors,
                    'signup_users': signup_users,
                    'conversion_rate': conversion_rate,
                    'contribution': contribution
                })
            
            # 가입자 수 기준으로 정렬
            core_data.sort(key=lambda x: x['signup_users'], reverse=True)
            analysis["core_funnel_analysis"] = core_data
        
        return analysis
    
    def _create_excel_report(self, analysis, filename="signup_path_analysis.xlsx"):
        """엑셀 파일로 분석 결과 출력 (핵심 지표만)"""
        print(f"📊 엑셀 리포트 생성 중: {filename}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "가입경로분석"
        
        # 스타일 정의
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        section_font = Font(bold=True, size=14)
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        current_row = 1
        
        # 1. 제목
        ws.cell(row=current_row, column=1, value="가입 경로 분석 (핵심 지표)")
        ws.cell(row=current_row, column=1).font = section_font
        current_row += 2
        
        # 2. 기본 정보
        ws.cell(row=current_row, column=1, value="분석 기간:")
        ws.cell(row=current_row, column=2, value="최근 30일")
        current_row += 1
        ws.cell(row=current_row, column=1, value="생성 일시:")
        ws.cell(row=current_row, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        current_row += 3
        
        # 3. 핵심 지표 표
        if analysis.get("core_funnel_analysis"):
            # 헤더
            headers = ['순위', '트래픽소스', '전체방문자', '가입완료자', '전환율(%)', '기여도(%)']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            current_row += 1
            
            # 데이터
            for i, data in enumerate(analysis["core_funnel_analysis"][:20], 1):
                ws.cell(row=current_row, column=1, value=i)
                ws.cell(row=current_row, column=2, value=data['source_medium'])
                ws.cell(row=current_row, column=3, value=data['all_visitors'])
                ws.cell(row=current_row, column=4, value=data['signup_users'])
                ws.cell(row=current_row, column=5, value=data['conversion_rate'])
                ws.cell(row=current_row, column=6, value=data['contribution'])
                current_row += 1
            
            current_row += 2
            
            # 4. 요약 정보
            ws.cell(row=current_row, column=1, value="📋 주요 인사이트")
            ws.cell(row=current_row, column=1).font = section_font
            current_row += 1
            
            # 총 가입자 수
            total_signups = sum([d['signup_users'] for d in analysis["core_funnel_analysis"]])
            ws.cell(row=current_row, column=1, value="총 가입자 수:")
            ws.cell(row=current_row, column=2, value=f"{total_signups:,}명")
            current_row += 1
            
            # 최고 볼륨 소스
            top_volume = analysis["core_funnel_analysis"][0]
            ws.cell(row=current_row, column=1, value="가장 많은 가입자를 유도한 소스:")
            ws.cell(row=current_row, column=2, value=f"{top_volume['source_medium']} ({top_volume['signup_users']}명)")
            current_row += 1
            
            # 최고 전환율 소스
            best_conversion = max(analysis["core_funnel_analysis"], key=lambda x: x['conversion_rate'])
            ws.cell(row=current_row, column=1, value="가장 높은 전환율을 보인 소스:")
            ws.cell(row=current_row, column=2, value=f"{best_conversion['source_medium']} ({best_conversion['conversion_rate']}%)")
            current_row += 1
            
            # 평균 전환율
            avg_conversion = sum([d['conversion_rate'] for d in analysis["core_funnel_analysis"]]) / len(analysis["core_funnel_analysis"])
            ws.cell(row=current_row, column=1, value="평균 전환율:")
            ws.cell(row=current_row, column=2, value=f"{round(avg_conversion, 1)}%")
        
        # 열 너비 자동 조정
        column_widths = [8, 45, 15, 15, 12, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        # 파일 저장
        wb.save(filename)
        print(f"✅ 엑셀 파일 생성 완료: {filename}")
        
        return filename
    
    def generate_report(self, analysis_result, filename="signup_path_analysis.json"):
        """분석 결과 리포트 생성"""
        report = {
            "generated_at": datetime.now().isoformat(),
            "analysis_type": "core_signup_funnel_analysis",
            "summary": {
                "total_sources": len(analysis_result.get("core_funnel_analysis", [])),
                "total_signups": sum([d['signup_users'] for d in analysis_result.get("core_funnel_analysis", [])]),
                "top_signup_source": analysis_result["core_funnel_analysis"][0]["source_medium"] if analysis_result.get("core_funnel_analysis") else "N/A",
                "highest_conversion_source": max(analysis_result.get("core_funnel_analysis", []), key=lambda x: x['conversion_rate'])["source_medium"] if analysis_result.get("core_funnel_analysis") else "N/A"
            },
            "detailed_analysis": analysis_result
        }
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
        
        print(f"✅ 분석 리포트 생성 완료: {filename}")
        return report
    
    def print_summary(self, analysis_result):
        """분석 결과 요약 출력"""
        print("\n" + "="*80)
        print("📊 가입 경로 핵심 지표 분석 결과")
        print("="*80)
        
        if analysis_result.get("core_funnel_analysis"):
            data = analysis_result["core_funnel_analysis"]
            
            print(f"\n🏆 상위 5개 가입 소스:")
            for i, source in enumerate(data[:5], 1):
                print(f"  {i}. {source['source_medium']}: {source['signup_users']}명 (전환율 {source['conversion_rate']}%)")
            
            print(f"\n🎯 최고 전환율 소스:")
            best_conversion = max(data, key=lambda x: x['conversion_rate'])
            print(f"  {best_conversion['source_medium']}: {best_conversion['conversion_rate']}%")
            
            print(f"\n📈 총 가입자 수: {sum([d['signup_users'] for d in data]):,}명")
            
            avg_conversion = sum([d['conversion_rate'] for d in data]) / len(data)
            print(f"📊 평균 전환율: {round(avg_conversion, 1)}%")


# 실행 예시
if __name__ == "__main__":
    # 설정 - 현재 디렉토리의 파일 경로로 수정
    CREDENTIALS_PATH = "/Users/brich/secure_keys/b-flow-407402-f835e3a78bf7.json"
    PROPERTY_ID = "302932513"
    
    # 분석기 초기화
    analyzer = UserSignupPathAnalyzer(CREDENTIALS_PATH, PROPERTY_ID)
    
    # 가입 경로 분석 실행
    print("🚀 실제 유저 기준 가입 퍼널 분석 시작...")
    analysis_result = analyzer.analyze_signup_sources(days=30)
    
    # 결과 출력
    analyzer.print_summary(analysis_result)
    
    # 리포트 생성
    report = analyzer.generate_report(analysis_result, "core_signup_funnel_analysis.json")
    
    print("\n🎉 분석 완료!")
    print("생성된 파일:")
    print("- signup_path_analysis.xlsx (핵심 지표 분석 리포트)")
    print("- core_signup_funnel_analysis.json (상세 분석 결과)")
    print("\n📊 깔끔하게 정리된 핵심 지표를 확인하세요!")