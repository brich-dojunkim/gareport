"""
리포트 생성 모듈
분석 결과를 다양한 형태의 리포트로 생성 (JSON, Excel, HTML)
"""

import json
import pandas as pd
import os
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from config import Config

class ReportGenerator:
    """리포트 생성 클래스"""
    
    def __init__(self, config: Config = None):
        """
        Args:
            config: 설정 객체
        """
        self.config = config or Config()
        
        # Excel 스타일 정의
        self.styles = {
            'header': {
                'font': Font(bold=True, color="FFFFFF", size=12),
                'fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
                'alignment': Alignment(horizontal="center", vertical="center")
            },
            'subheader': {
                'font': Font(bold=True, size=11),
                'fill': PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
                'alignment': Alignment(horizontal="center", vertical="center")
            },
            'title': {
                'font': Font(bold=True, size=16),
                'alignment': Alignment(horizontal="center", vertical="center")
            },
            'metric': {
                'font': Font(bold=True, size=14, color="2F75B5"),
                'alignment': Alignment(horizontal="center", vertical="center")
            },
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    def generate_optimization_report(self, conversion_rates: Dict, 
                                   patterns: Dict, output_dir: str) -> Dict:
        """
        종합 최적화 리포트 생성
        
        Args:
            conversion_rates: 전환율 데이터
            patterns: 패턴 분석 결과
            output_dir: 출력 디렉토리
            
        Returns:
            생성된 리포트 파일 정보
        """
        print("📋 최적화 리포트 생성 중...")
        
        # 출력 디렉토리 생성
        os.makedirs(output_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        report_files = {}
        
        # 1. JSON 리포트 생성
        if 'json' in self.config.reporting['output_formats']:
            json_file = self._generate_json_report(
                conversion_rates, patterns, 
                os.path.join(output_dir, f'funnel_analysis_{timestamp}.json')
            )
            report_files['json'] = json_file
        
        # 2. Excel 리포트 생성
        if 'excel' in self.config.reporting['output_formats']:
            excel_file = self._generate_excel_report(
                conversion_rates, patterns,
                os.path.join(output_dir, f'funnel_analysis_{timestamp}.xlsx')
            )
            report_files['excel'] = excel_file
        
        # 3. HTML 리포트 생성
        if 'html' in self.config.reporting['output_formats']:
            html_file = self._generate_html_report(
                conversion_rates, patterns,
                os.path.join(output_dir, f'funnel_analysis_{timestamp}.html')
            )
            report_files['html'] = html_file
        
        # 4. 요약 리포트 생성
        summary_file = self._generate_summary_report(
            conversion_rates, patterns,
            os.path.join(output_dir, f'executive_summary_{timestamp}.json')
        )
        report_files['summary'] = summary_file
        
        print(f"✅ 리포트 생성 완료: {len(report_files)}개 파일")
        for format_type, file_path in report_files.items():
            print(f"  - {format_type.upper()}: {file_path}")
        
        return report_files
    
    def _generate_json_report(self, conversion_rates: Dict, patterns: Dict, 
                            file_path: str) -> str:
        """JSON 형태의 상세 리포트 생성"""
        print("  📄 JSON 리포트 생성...")
        
        report_data = {
            'report_metadata': {
                'generated_at': datetime.now().isoformat(),
                'report_type': 'B-flow_funnel_analysis',
                'version': '1.0.0',
                'analysis_period': 'Last 30 days'
            },
            'executive_summary': {
                'total_users': conversion_rates.get('total_users', 0),
                'overall_conversion_rate': conversion_rates.get('overall_conversion', 0),
                'key_insights': self._extract_key_insights(conversion_rates, patterns),
                'top_recommendations': self._get_top_recommendations(patterns)
            },
            'funnel_performance': {
                'stage_conversion_rates': conversion_rates.get('step_conversions', {}),
                'stage_counts': conversion_rates.get('stage_counts', {}),
                'bottleneck_analysis': self._identify_bottleneck(conversion_rates)
            },
            'traffic_analysis': {
                'source_performance': patterns.get('source_page_combinations', []),
                'high_value_segments': patterns.get('high_value_segments', []),
                'optimization_opportunities': patterns.get('optimization_insights', {})
            },
            'user_journey_insights': {
                'journey_patterns': patterns.get('journey_patterns', {}),
                'engagement_patterns': patterns.get('engagement_patterns', {}),
                'drop_off_analysis': patterns.get('drop_off_points', [])
            },
            'detailed_data': {
                'conversion_rates': conversion_rates,
                'patterns': patterns
            }
        }
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, ensure_ascii=False, indent=2, default=str)
        
        return file_path
    
    def _generate_excel_report(self, conversion_rates: Dict, patterns: Dict,
                             file_path: str) -> str:
        """Excel 형태의 시각적 리포트 생성"""
        print("  📊 Excel 리포트 생성...")
        
        wb = Workbook()
        
        # 기본 워크시트 제거
        wb.remove(wb.active)
        
        # 시트 생성
        self._create_summary_sheet(wb, conversion_rates, patterns)
        self._create_funnel_analysis_sheet(wb, conversion_rates)
        self._create_traffic_analysis_sheet(wb, patterns)
        self._create_journey_analysis_sheet(wb, patterns)
        self._create_optimization_sheet(wb, patterns)
        
        wb.save(file_path)
        return file_path
    
    def _create_summary_sheet(self, wb: Workbook, conversion_rates: Dict, patterns: Dict):
        """요약 시트 생성"""
        ws = wb.create_sheet("📊 Executive Summary")
        
        # 제목
        ws['B2'] = 'B-flow 퍼널 분석 요약 리포트'
        ws['B2'].font = self.styles['title']['font']
        ws['B2'].alignment = self.styles['title']['alignment']
        ws.merge_cells('B2:H2')
        
        # 생성 일시
        ws['B3'] = f'생성일시: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws.merge_cells('B3:H3')
        
        # 주요 지표
        current_row = 5
        ws[f'B{current_row}'] = '🎯 핵심 성과 지표'
        ws[f'B{current_row}'].font = self.styles['subheader']['font']
        ws[f'B{current_row}'].fill = self.styles['subheader']['fill']
        ws.merge_cells(f'B{current_row}:H{current_row}')
        
        current_row += 2
        metrics = [
            ('전체 사용자 수', f"{conversion_rates.get('total_users', 0):,}명"),
            ('전체 전환율', f"{conversion_rates.get('overall_conversion', 0):.1f}%"),
            ('Awareness → Interest', f"{conversion_rates.get('step_conversions', {}).get('awareness_to_interest', 0):.1f}%"),
            ('Interest → Consideration', f"{conversion_rates.get('step_conversions', {}).get('interest_to_consideration', 0):.1f}%"),
            ('Consideration → Conversion', f"{conversion_rates.get('step_conversions', {}).get('consideration_to_conversion', 0):.1f}%")
        ]
        
        for metric_name, metric_value in metrics:
            ws[f'C{current_row}'] = metric_name
            ws[f'F{current_row}'] = metric_value
            ws[f'F{current_row}'].font = self.styles['metric']['font']
            current_row += 1
        
        # 주요 인사이트
        current_row += 2
        ws[f'B{current_row}'] = '💡 주요 인사이트'
        ws[f'B{current_row}'].font = self.styles['subheader']['font']
        ws[f'B{current_row}'].fill = self.styles['subheader']['fill']
        ws.merge_cells(f'B{current_row}:H{current_row}')
        
        current_row += 2
        insights = self._extract_key_insights(conversion_rates, patterns)
        for i, insight in enumerate(insights[:5], 1):
            ws[f'C{current_row}'] = f"{i}. {insight}"
            current_row += 1
        
        # 최우선 권장사항
        current_row += 2
        ws[f'B{current_row}'] = '🚀 최우선 권장사항'
        ws[f'B{current_row}'].font = self.styles['subheader']['font']
        ws[f'B{current_row}'].fill = self.styles['subheader']['fill']
        ws.merge_cells(f'B{current_row}:H{current_row}')
        
        current_row += 2
        recommendations = self._get_top_recommendations(patterns)
        for i, rec in enumerate(recommendations[:3], 1):
            ws[f'C{current_row}'] = f"{i}. {rec}"
            current_row += 1
        
        # 열 너비 조정
        column_widths = [3, 25, 5, 5, 20, 10, 3]
        for i, width in enumerate(column_widths, 2):
            ws.column_dimensions[chr(64 + i)].width = width
    
    def _create_funnel_analysis_sheet(self, wb: Workbook, conversion_rates: Dict):
        """퍼널 분석 시트 생성"""
        ws = wb.create_sheet("🔄 Funnel Analysis")
        
        # 제목
        ws['B2'] = '퍼널 단계별 전환 분석'
        ws['B2'].font = self.styles['title']['font']
        ws.merge_cells('B2:G2')
        
        # 테이블 헤더
        current_row = 4
        headers = ['퍼널 단계', '사용자 수', '전환율 (%)', '이전 단계 대비 (%)', '개선 우선순위']
        
        for col, header in enumerate(headers, 2):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = self.styles['header']['font']
            cell.fill = self.styles['header']['fill']
            cell.alignment = self.styles['header']['alignment']
            cell.border = self.styles['border']
        
        current_row += 1
        
        # 단계별 데이터
        stage_data = [
            ('AWARENESS', 
             conversion_rates.get('stage_counts', {}).get('awareness', 0), 
             conversion_rates.get('stage_rates', {}).get('awareness_rate', 0), 
             '-', 'Low'),
            ('INTEREST', 
             conversion_rates.get('stage_counts', {}).get('interest', 0),
             conversion_rates.get('stage_rates', {}).get('interest_rate', 0),
             conversion_rates.get('step_conversions', {}).get('awareness_to_interest', 0), 
             'Medium'),
            ('CONSIDERATION', 
             conversion_rates.get('stage_counts', {}).get('consideration', 0),
             conversion_rates.get('stage_rates', {}).get('consideration_rate', 0),
             conversion_rates.get('step_conversions', {}).get('interest_to_consideration', 0), 
             'High'),
            ('CONVERSION', 
             conversion_rates.get('stage_counts', {}).get('conversion', 0),
             conversion_rates.get('stage_rates', {}).get('conversion_rate', 0),
             conversion_rates.get('step_conversions', {}).get('consideration_to_conversion', 0), 
             'Critical')
        ]
        
        for stage_name, users, rate, step_rate, priority in stage_data:
            ws.cell(row=current_row, column=2, value=stage_name).border = self.styles['border']
            ws.cell(row=current_row, column=3, value=f"{users:,}").border = self.styles['border']
            ws.cell(row=current_row, column=4, value=f"{rate:.1f}%").border = self.styles['border']
            
            if step_rate != '-':
                ws.cell(row=current_row, column=5, value=f"{step_rate:.1f}%").border = self.styles['border']
            else:
                ws.cell(row=current_row, column=5, value=step_rate).border = self.styles['border']
            
            ws.cell(row=current_row, column=6, value=priority).border = self.styles['border']
            current_row += 1
        
        # 열 너비 조정
        column_widths = [3, 15, 12, 12, 18, 15]
        for i, width in enumerate(column_widths, 2):
            ws.column_dimensions[chr(64 + i)].width = width
    
    def _create_traffic_analysis_sheet(self, wb: Workbook, patterns: Dict):
        """트래픽 분석 시트 생성"""
        ws = wb.create_sheet("🚀 Traffic Analysis")
        
        # 제목
        ws['B2'] = '트래픽 소스별 성과 분석'
        ws['B2'].font = self.styles['title']['font']
        ws.merge_cells('B2:H2')
        
        # 소스별 성과 테이블
        current_row = 4
        source_patterns = patterns.get('source_page_combinations', [])
        
        if source_patterns:
            headers = ['순위', '트래픽 그룹', '최적 조합', '전환율 (%)', '사용자 수', '권장 액션']
            
            for col, header in enumerate(headers, 2):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = self.styles['header']['font']
                cell.fill = self.styles['header']['fill']
                cell.alignment = self.styles['header']['alignment']
                cell.border = self.styles['border']
            
            current_row += 1
            
            for rank, pattern in enumerate(source_patterns[:10], 1):
                best_combo = pattern['best_combination']
                ws.cell(row=current_row, column=2, value=rank).border = self.styles['border']
                ws.cell(row=current_row, column=3, value=pattern['traffic_group']).border = self.styles['border']
                ws.cell(row=current_row, column=4, value=best_combo['combination']).border = self.styles['border']
                ws.cell(row=current_row, column=5, value=f"{best_combo['conversion_rate']:.1f}%").border = self.styles['border']
                ws.cell(row=current_row, column=6, value=f"{best_combo['users']:,}").border = self.styles['border']
                
                # 권장 액션
                if best_combo['conversion_rate'] > 20:
                    action = "예산 확대"
                elif best_combo['conversion_rate'] > 10:
                    action = "테스트 확대"
                else:
                    action = "최적화 필요"
                ws.cell(row=current_row, column=7, value=action).border = self.styles['border']
                current_row += 1
        
        # 고가치 세그먼트
        current_row += 3
        ws[f'B{current_row}'] = '💎 고가치 사용자 세그먼트'
        ws[f'B{current_row}'].font = self.styles['subheader']['font']
        ws[f'B{current_row}'].fill = self.styles['subheader']['fill']
        ws.merge_cells(f'B{current_row}:H{current_row}')
        
        current_row += 2
        high_value_segments = patterns.get('high_value_segments', [])
        
        if high_value_segments:
            headers = ['세그먼트명', '설명', '사용자 수', '전환율 (%)', '가치 점수', '타겟팅 전략']
            
            for col, header in enumerate(headers, 2):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = self.styles['header']['font']
                cell.fill = self.styles['header']['fill']
                cell.alignment = self.styles['header']['alignment']
                cell.border = self.styles['border']
            
            current_row += 1
            
            for segment in high_value_segments[:5]:
                ws.cell(row=current_row, column=2, value=segment['segment_name']).border = self.styles['border']
                ws.cell(row=current_row, column=3, value=segment['description']).border = self.styles['border']
                ws.cell(row=current_row, column=4, value=f"{segment['users_count']:,}").border = self.styles['border']
                ws.cell(row=current_row, column=5, value=f"{segment['conversion_rate']:.1f}%").border = self.styles['border']
                ws.cell(row=current_row, column=6, value=f"{segment['segment_value']:.0f}").border = self.styles['border']
                ws.cell(row=current_row, column=7, value="맞춤형 캠페인").border = self.styles['border']
                current_row += 1
        
        # 열 너비 조정
        column_widths = [3, 15, 20, 15, 12, 12, 15]
        for i, width in enumerate(column_widths, 2):
            ws.column_dimensions[chr(64 + i)].width = width
    
    def _create_journey_analysis_sheet(self, wb: Workbook, patterns: Dict):
        """사용자 여정 분석 시트 생성"""
        ws = wb.create_sheet("🛤️ User Journey")
        
        # 제목
        ws['B2'] = '사용자 여정 패턴 분석'
        ws['B2'].font = self.styles['title']['font']
        ws.merge_cells('B2:H2')
        
        # 고전환 여정 패턴
        current_row = 4
        ws[f'B{current_row}'] = '🎯 고전환 여정 패턴'
        ws[f'B{current_row}'].font = self.styles['subheader']['font']
        ws[f'B{current_row}'].fill = self.styles['subheader']['fill']
        ws.merge_cells(f'B{current_row}:H{current_row}')
        
        current_row += 2
        journey_patterns = patterns.get('journey_patterns', {})
        common_paths = journey_patterns.get('common_paths', [])
        
        if common_paths:
            headers = ['순위', '여정 경로', '사용자 수', '전환 수', '전환율 (%)', '최적화 기회']
            
            for col, header in enumerate(headers, 2):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = self.styles['header']['font']
                cell.fill = self.styles['header']['fill']
                cell.alignment = self.styles['header']['alignment']
                cell.border = self.styles['border']
            
            current_row += 1
            
            for rank, path in enumerate(common_paths[:10], 1):
                ws.cell(row=current_row, column=2, value=rank).border = self.styles['border']
                ws.cell(row=current_row, column=3, value=path['path']).border = self.styles['border']
                ws.cell(row=current_row, column=4, value=f"{path['users']:,}").border = self.styles['border']
                ws.cell(row=current_row, column=5, value=f"{path['conversions']:,}").border = self.styles['border']
                ws.cell(row=current_row, column=6, value=f"{path['conversion_rate']:.1f}%").border = self.styles['border']
                
                # 최적화 기회
                if path['conversion_rate'] > 50:
                    opportunity = "패턴 확산"
                elif path['conversion_rate'] > 20:
                    opportunity = "경험 개선"
                else:
                    opportunity = "경로 최적화"
                ws.cell(row=current_row, column=7, value=opportunity).border = self.styles['border']
                current_row += 1
        
        # 이탈 지점 분석
        current_row += 3
        ws[f'B{current_row}'] = '🚪 주요 이탈 지점'
        ws[f'B{current_row}'].font = self.styles['subheader']['font']
        ws[f'B{current_row}'].fill = self.styles['subheader']['fill']
        ws.merge_cells(f'B{current_row}:H{current_row}')
        
        current_row += 2
        drop_off_points = patterns.get('drop_off_points', [])
        
        if drop_off_points:
            headers = ['단계', '이탈자 수', '이탈률 (%)', '주요 이탈 페이지', '개선 우선순위']
            
            for col, header in enumerate(headers, 2):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = self.styles['header']['font']
                cell.fill = self.styles['header']['fill']
                cell.alignment = self.styles['header']['alignment']
                cell.border = self.styles['border']
            
            current_row += 1
            
            for dropoff in drop_off_points:
                ws.cell(row=current_row, column=2, value=dropoff['stage']).border = self.styles['border']
                ws.cell(row=current_row, column=3, value=f"{dropoff['total_dropoffs']:,}").border = self.styles['border']
                ws.cell(row=current_row, column=4, value=f"{dropoff['dropoff_rate']:.1f}%").border = self.styles['border']
                
                if dropoff['top_dropoff_pages']:
                    top_page = dropoff['top_dropoff_pages'][0]['page']
                    ws.cell(row=current_row, column=5, value=top_page).border = self.styles['border']
                
                # 우선순위
                if dropoff['dropoff_rate'] > 30:
                    priority = "긴급"
                elif dropoff['dropoff_rate'] > 20:
                    priority = "높음"
                else:
                    priority = "보통"
                ws.cell(row=current_row, column=6, value=priority).border = self.styles['border']
                current_row += 1
        
        # 열 너비 조정
        column_widths = [3, 25, 12, 12, 12, 15]
        for i, width in enumerate(column_widths, 2):
            ws.column_dimensions[chr(64 + i)].width = width
    
    def _create_optimization_sheet(self, wb: Workbook, patterns: Dict):
        """최적화 권장사항 시트 생성"""
        ws = wb.create_sheet("🚀 Optimization")
        
        # 제목
        ws['B2'] = '최적화 권장사항 및 액션 플랜'
        ws['B2'].font = self.styles['title']['font']
        ws.merge_cells('B2:I2')
        
        # 최우선 기회
        current_row = 4
        ws[f'B{current_row}'] = '🎯 최우선 최적화 기회'
        ws[f'B{current_row}'].font = self.styles['subheader']['font']
        ws[f'B{current_row}'].fill = self.styles['subheader']['fill']
        ws.merge_cells(f'B{current_row}:I{current_row}')
        
        current_row += 2
        optimization_insights = patterns.get('optimization_insights', {})
        top_opportunities = optimization_insights.get('top_opportunities', [])
        
        if top_opportunities:
            headers = ['우선순위', '최적화 유형', '대상', '현재 성과', '권장 액션', '예상 효과', '구현 난이도', '담당팀']
            
            for col, header in enumerate(headers, 2):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = self.styles['header']['font']
                cell.fill = self.styles['header']['fill']
                cell.alignment = self.styles['header']['alignment']
                cell.border = self.styles['border']
            
            current_row += 1
            
            for opportunity in top_opportunities:
                ws.cell(row=current_row, column=2, value=opportunity['priority']).border = self.styles['border']
                ws.cell(row=current_row, column=3, value=opportunity['type']).border = self.styles['border']
                
                # 대상
                target = opportunity.get('segment', opportunity.get('traffic_group', opportunity.get('category', 'N/A')))
                ws.cell(row=current_row, column=4, value=target).border = self.styles['border']
                
                # 현재 성과
                current_performance = f"{opportunity.get('conversion_rate', 0):.1f}%" if 'conversion_rate' in opportunity else 'N/A'
                ws.cell(row=current_row, column=5, value=current_performance).border = self.styles['border']
                
                ws.cell(row=current_row, column=6, value=opportunity['recommendation']).border = self.styles['border']
                
                # 예상 효과
                if opportunity['priority'] == 'High':
                    expected_effect = "20-50% 개선"
                elif opportunity['priority'] == 'Medium':
                    expected_effect = "10-30% 개선"
                else:
                    expected_effect = "5-15% 개선"
                ws.cell(row=current_row, column=7, value=expected_effect).border = self.styles['border']
                
                # 구현 난이도
                if 'Scale' in opportunity['type'] or 'Target' in opportunity['type']:
                    difficulty = "쉬움"
                elif 'Optimize' in opportunity['type']:
                    difficulty = "보통"
                else:
                    difficulty = "어려움"
                ws.cell(row=current_row, column=8, value=difficulty).border = self.styles['border']
                
                # 담당팀
                if 'Traffic' in opportunity['type'] or 'Scale' in opportunity['type']:
                    team = "마케팅팀"
                elif 'Content' in opportunity['type']:
                    team = "콘텐츠팀"
                else:
                    team = "제품팀"
                ws.cell(row=current_row, column=9, value=team).border = self.styles['border']
                
                current_row += 1
        
        # 실행 로드맵
        current_row += 3
        ws[f'B{current_row}'] = '📅 4주 실행 로드맵'
        ws[f'B{current_row}'].font = self.styles['subheader']['font']
        ws[f'B{current_row}'].fill = self.styles['subheader']['fill']
        ws.merge_cells(f'B{current_row}:F{current_row}')
        
        current_row += 2
        roadmap = [
            ('1주차', 'Quick Wins 실행', '고전환 세그먼트 타겟팅 강화', '마케팅팀'),
            ('2주차', '주요 이탈지점 개선', 'CONSIDERATION 단계 페이지 최적화', '제품팀'),
            ('3주차', '고성과 콘텐츠 확산', '성과 좋은 콘텐츠 유형 추가 제작', '콘텐츠팀'),
            ('4주차', '성과 측정 및 반복', '개선 효과 측정 및 다음 액션 계획', '전체팀')
        ]
        
        headers = ['기간', '액션 유형', '구체적 실행사항', '담당팀']
        for col, header in enumerate(headers, 2):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = self.styles['header']['font']
            cell.fill = self.styles['header']['fill']
            cell.alignment = self.styles['header']['alignment']
            cell.border = self.styles['border']
        
        current_row += 1
        
        for period, action_type, specific_action, team in roadmap:
            ws.cell(row=current_row, column=2, value=period).border = self.styles['border']
            ws.cell(row=current_row, column=3, value=action_type).border = self.styles['border']
            ws.cell(row=current_row, column=4, value=specific_action).border = self.styles['border']
            ws.cell(row=current_row, column=5, value=team).border = self.styles['border']
            current_row += 1
        
        # 열 너비 조정
        column_widths = [3, 12, 15, 30, 12, 12, 15, 12, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
    
    def _generate_html_report(self, conversion_rates: Dict, patterns: Dict,
                            file_path: str) -> str:
        """HTML 형태의 대시보드 리포트 생성"""
        print("  🌐 HTML 리포트 생성...")
        
        html_template = """
        <!DOCTYPE html>
        <html lang="ko">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>B-flow 퍼널 분석 리포트</title>
            <style>
                body {{ font-family: 'Arial', sans-serif; margin: 0; padding: 20px; background-color: #f5f5f5; }}
                .container {{ max-width: 1200px; margin: 0 auto; background-color: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                .header {{ text-align: center; margin-bottom: 40px; }}
                .header h1 {{ color: #2c3e50; margin-bottom: 10px; }}
                .header .subtitle {{ color: #7f8c8d; font-size: 16px; }}
                .metrics-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); gap: 20px; margin-bottom: 40px; }}
                .metric-card {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 25px; border-radius: 10px; text-align: center; }}
                .metric-card h3 {{ margin: 0 0 10px 0; font-size: 14px; opacity: 0.9; }}
                .metric-card .value {{ font-size: 32px; font-weight: bold; margin: 0; }}
                .section {{ margin-bottom: 40px; }}
                .section h2 {{ color: #2c3e50; border-bottom: 3px solid #3498db; padding-bottom: 10px; }}
                .table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
                .table th {{ background-color: #34495e; color: white; padding: 12px; text-align: left; }}
                .table td {{ padding: 12px; border-bottom: 1px solid #ecf0f1; }}
                .table tr:hover {{ background-color: #f8f9fa; }}
                .priority-high {{ background-color: #e74c3c; color: white; padding: 4px 8px; border-radius: 4px; }}
                .priority-medium {{ background-color: #f39c12; color: white; padding: 4px 8px; border-radius: 4px; }}
                .priority-low {{ background-color: #27ae60; color: white; padding: 4px 8px; border-radius: 4px; }}
                .insight-box {{ background-color: #ecf0f1; padding: 20px; border-radius: 8px; margin: 20px 0; border-left: 4px solid #3498db; }}
                .recommendation {{ background-color: #d5f4e6; padding: 15px; border-radius: 8px; margin: 10px 0; border-left: 4px solid #27ae60; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>🚀 B-flow 퍼널 분석 리포트</h1>
                    <div class="subtitle">생성일시: {generated_at}</div>
                </div>
                
                <div class="metrics-grid">
                    {metrics_cards}
                </div>
                
                <div class="section">
                    <h2>📊 퍼널 성과 개요</h2>
                    {funnel_overview}
                </div>
                
                <div class="section">
                    <h2>🎯 최우선 최적화 기회</h2>
                    {optimization_opportunities}
                </div>
                
                <div class="section">
                    <h2>🚀 권장 액션 플랜</h2>
                    {action_plan}
                </div>
            </div>
        </body>
        </html>
        """
        
        # HTML 구성 요소 생성
        metrics_cards = self._generate_metrics_cards_html(conversion_rates)
        funnel_overview = self._generate_funnel_overview_html(conversion_rates)
        optimization_opportunities = self._generate_optimization_html(patterns)
        action_plan = self._generate_action_plan_html()
        
        # HTML 생성
        html_content = html_template.format(
            generated_at=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            metrics_cards=metrics_cards,
            funnel_overview=funnel_overview,
            optimization_opportunities=optimization_opportunities,
            action_plan=action_plan
        )
        
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        return file_path
    
    def _generate_summary_report(self, conversion_rates: Dict, patterns: Dict,
                               file_path: str) -> str:
        """경영진용 요약 리포트 생성"""
        print("  📋 경영진 요약 리포트 생성...")
        
        summary = {
            'executive_summary': {
                'report_date': datetime.now().strftime('%Y-%m-%d'),
                'analysis_period': 'Last 30 days',
                'total_users': conversion_rates.get('total_users', 0),
                'conversion_rate': conversion_rates.get('overall_conversion', 0),
                'key_findings': self._extract_key_insights(conversion_rates, patterns)[:3],
                'immediate_actions': self._get_immediate_actions(patterns)[:3],
                'expected_impact': self._calculate_expected_impact(conversion_rates, patterns)
            },
            'performance_snapshot': {
                'funnel_health': self._assess_funnel_health(conversion_rates),
                'traffic_quality': self._assess_traffic_quality(patterns),
                'optimization_score': self._calculate_optimization_score(patterns)
            },
            'recommendations': {
                'week_1': '고전환 세그먼트 타겟팅 집중',
                'week_2': '주요 이탈지점 UX 개선',
                'week_3': '성과 콘텐츠 스케일링',
                'week_4': '성과 측정 및 반복 최적화'
            }
        }
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(summary, f, ensure_ascii=False, indent=2, default=str)
        
        return file_path
    
    # Helper 메서드들
    def _extract_key_insights(self, conversion_rates: Dict, patterns: Dict) -> List[str]:
        """주요 인사이트 추출"""
        insights = []
        
        # 전환율 분석
        overall_rate = conversion_rates.get('overall_conversion', 0)
        if overall_rate > 10:
            insights.append(f"전체 전환율 {overall_rate:.1f}%로 양호한 성과")
        elif overall_rate > 5:
            insights.append(f"전체 전환율 {overall_rate:.1f}%로 개선 여지 있음")
        else:
            insights.append(f"전체 전환율 {overall_rate:.1f}%로 시급한 개선 필요")
        
        # 병목 지점 식별
        bottleneck = self._identify_bottleneck(conversion_rates)
        if bottleneck:
            insights.append(f"주요 병목: {bottleneck['stage']} 단계 ({bottleneck['rate']:.1f}%)")
        
        # 고성과 세그먼트
        high_value_segments = patterns.get('high_value_segments', [])
        if high_value_segments:
            top_segment = high_value_segments[0]
            insights.append(f"고가치 세그먼트: {top_segment['segment_name']} ({top_segment['conversion_rate']:.1f}% 전환율)")
        
        return insights
    
    def _get_top_recommendations(self, patterns: Dict) -> List[str]:
        """주요 권장사항 추출"""
        recommendations = []
        
        optimization_insights = patterns.get('optimization_insights', {})
        top_opportunities = optimization_insights.get('top_opportunities', [])
        
        for opportunity in top_opportunities[:3]:
            recommendations.append(opportunity.get('recommendation', ''))
        
        return recommendations
    
    def _identify_bottleneck(self, conversion_rates: Dict) -> Optional[Dict]:
        """주요 병목 지점 식별"""
        step_conversions = conversion_rates.get('step_conversions', {})
        
        if not step_conversions:
            return None
        
        min_step = min(step_conversions.items(), key=lambda x: x[1])
        return {
            'stage': min_step[0],
            'rate': min_step[1]
        }
    
    def _get_immediate_actions(self, patterns: Dict) -> List[str]:
        """즉시 실행 가능한 액션들"""
        actions = []
        
        optimization_insights = patterns.get('optimization_insights', {})
        quick_wins = optimization_insights.get('quick_wins', [])
        
        for quick_win in quick_wins:
            actions.append(quick_win.get('recommendation', ''))
        
        return actions
    
    def _calculate_expected_impact(self, conversion_rates: Dict, patterns: Dict) -> str:
        """예상 효과 계산"""
        current_rate = conversion_rates.get('overall_conversion', 0)
        
        if current_rate > 10:
            return "15-25% 전환율 개선 기대"
        elif current_rate > 5:
            return "25-40% 전환율 개선 기대"
        else:
            return "50-100% 전환율 개선 기대"
    
    def _assess_funnel_health(self, conversion_rates: Dict) -> str:
        """퍼널 건강도 평가"""
        overall_rate = conversion_rates.get('overall_conversion', 0)
        
        if overall_rate > 15:
            return "Excellent"
        elif overall_rate > 10:
            return "Good"
        elif overall_rate > 5:
            return "Fair"
        else:
            return "Poor"
    
    def _assess_traffic_quality(self, patterns: Dict) -> str:
        """트래픽 품질 평가"""
        high_value_segments = patterns.get('high_value_segments', [])
        
        if len(high_value_segments) >= 3:
            return "High"
        elif len(high_value_segments) >= 1:
            return "Medium"
        else:
            return "Low"
    
    def _calculate_optimization_score(self, patterns: Dict) -> int:
        """최적화 점수 계산 (100점 만점)"""
        score = 50  # 기본 점수
        
        optimization_insights = patterns.get('optimization_insights', {})
        opportunities = optimization_insights.get('top_opportunities', [])
        
        # 기회의 개수와 우선순위에 따라 점수 조정
        for opportunity in opportunities[:5]:
            if opportunity.get('priority') == 'High':
                score += 15
            elif opportunity.get('priority') == 'Medium':
                score += 10
            else:
                score += 5
        
        return min(score, 100)
    
    def _generate_metrics_cards_html(self, conversion_rates: Dict) -> str:
        """메트릭 카드 HTML 생성"""
        cards = []
        
        metrics = [
            ('전체 사용자', f"{conversion_rates.get('total_users', 0):,}명"),
            ('전체 전환율', f"{conversion_rates.get('overall_conversion', 0):.1f}%"),
            ('Awareness→Interest', f"{conversion_rates.get('step_conversions', {}).get('awareness_to_interest', 0):.1f}%"),
            ('Interest→Consideration', f"{conversion_rates.get('step_conversions', {}).get('interest_to_consideration', 0):.1f}%"),
            ('Consideration→Conversion', f"{conversion_rates.get('step_conversions', {}).get('consideration_to_conversion', 0):.1f}%")
        ]
        
        for title, value in metrics:
            cards.append(f'<div class="metric-card"><h3>{title}</h3><div class="value">{value}</div></div>')
        
        return '\n'.join(cards)
    
    def _generate_funnel_overview_html(self, conversion_rates: Dict) -> str:
        """퍼널 개요 HTML 생성"""
        bottleneck = self._identify_bottleneck(conversion_rates)
        bottleneck_text = f"주요 병목: {bottleneck['stage']} 단계" if bottleneck else "병목 지점 없음"
        
        return f'''
        <div class="insight-box">
            <strong>퍼널 건강도:</strong> {self._assess_funnel_health(conversion_rates)}<br>
            <strong>{bottleneck_text}</strong><br>
            <strong>총 전환 수:</strong> {conversion_rates.get('stage_counts', {}).get('conversion', 0):,}건
        </div>
        '''
    
    def _generate_optimization_html(self, patterns: Dict) -> str:
        """최적화 기회 HTML 생성"""
        optimization_insights = patterns.get('optimization_insights', {})
        opportunities = optimization_insights.get('top_opportunities', [])[:3]
        
        html_parts = []
        for i, opportunity in enumerate(opportunities, 1):
            priority_class = f"priority-{opportunity.get('priority', 'low').lower()}"
            html_parts.append(f'''
            <div class="recommendation">
                <strong>{i}. {opportunity.get('type', '')}</strong>
                <span class="{priority_class}">{opportunity.get('priority', '')}</span><br>
                {opportunity.get('recommendation', '')}
            </div>
            ''')
        
        return '\n'.join(html_parts)
    
    def _generate_action_plan_html(self) -> str:
        """액션 플랜 HTML 생성"""
        return '''
        <div class="insight-box">
            <h3>📅 4주 실행 로드맵</h3>
            <ul>
                <li><strong>1주차:</strong> 고전환 세그먼트 타겟팅 집중 (마케팅팀)</li>
                <li><strong>2주차:</strong> 주요 이탈지점 UX 개선 (제품팀)</li>
                <li><strong>3주차:</strong> 성과 콘텐츠 스케일링 (콘텐츠팀)</li>
                <li><strong>4주차:</strong> 성과 측정 및 반복 최적화 (전체팀)</li>
            </ul>
        </div>
        '''

# 사용 예시
if __name__ == "__main__":
    # 설정 및 리포트 생성기 초기화
    config = Config()
    generator = ReportGenerator(config)
    
    # 샘플 데이터
    sample_conversion_rates = {
        'total_users': 1000,
        'overall_conversion': 12.5,
        'stage_counts': {
            'awareness': 1000,
            'interest': 650,
            'consideration': 300,
            'conversion': 125
        },
        'stage_rates': {
            'awareness_rate': 100.0,
            'interest_rate': 65.0,
            'consideration_rate': 30.0,
            'conversion_rate': 12.5
        },
        'step_conversions': {
            'awareness_to_interest': 65.0,
            'interest_to_consideration': 46.2,
            'consideration_to_conversion': 41.7
        }
    }
    
    sample_patterns = {
        'optimization_insights': {
            'top_opportunities': [
                {
                    'type': 'Scale High Performer',
                    'priority': 'High',
                    'recommendation': 'google/cpc 트래픽 예산 확대',
                    'conversion_rate': 25.0,
                    'traffic_group': 'paid_search'
                }
            ],
            'quick_wins': [
                {
                    'type': 'Target High-Value Segment',
                    'recommendation': '다중 세션 사용자 타겟팅 강화',
                    'priority': 'High'
                }
            ]
        },
        'high_value_segments': [
            {
                'segment_name': 'Content Engaged Users',
                'description': '콘텐츠 참여 후 서비스 탐색 사용자',
                'conversion_rate': 35.0,
                'users_count': 200,
                'segment_value': 7000
            }
        ],
        'source_page_combinations': [
            {
                'traffic_group': 'paid_search',
                'best_combination': {
                    'combination': 'homepage + content + service_info',
                    'users': 300,
                    'conversions': 75,
                    'conversion_rate': 25.0
                }
            }
        ],
        'journey_patterns': {
            'common_paths': [
                {
                    'path': '/ → /posts/739 → /provider-join',
                    'users': 150,
                    'conversions': 90,
                    'conversion_rate': 60.0
                }
            ]
        },
        'drop_off_points': [
            {
                'stage': 'CONSIDERATION',
                'total_dropoffs': 200,
                'dropoff_rate': 35.0,
                'top_dropoff_pages': [
                    {'page': '/fee-information', 'count': 80, 'percentage': 40.0}
                ]
            }
        ]
    }
    
    # 리포트 생성
    try:
        report_files = generator.generate_optimization_report(
            sample_conversion_rates, sample_patterns, 'test_output'
        )
        print(f"\n✅ 리포트 생성 완료: {len(report_files)}개 파일")
        for format_type, file_path in report_files.items():
            print(f"  - {format_type.upper()}: {file_path}")
            
    except Exception as e:
        print(f"❌ 오류 발생: {str(e)}")
        import traceback
        traceback.print_exc()